# -*- coding: utf-8 -*-
"""Export active vacancies for the public search site.

PUBLIC RULE: never export заказчик / адрес проживания / адрес объекта
into fields that the website shows to candidates or partners.

Card rule: each должность with an explicit demand digit becomes its own card.
Rows without digits (legacy) stay as one combined card until the sheet is fixed.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
from geo import geocode, load_cache, save_cache  # noqa: E402

XLSX = ROOT / "data" / "Для сотрудников подбора.xlsx"
OUT = ROOT / "vacancies.json"

FEM_HINTS = (
    "щица",
    "чица",
    "ница",
    "варка",
    "торка",
    "истка",
    "есса",
    "уборк",
    "горнич",
    "швея",
    "посудомо",
    "фасовщи",
    "упаковщиц",
    "комплектовщиц",
    "маркировщиц",
    "маркеровщиц",
    "сканировщиц",
    "сортировщиц",
)
MASC_HINTS = (
    "грузчик",
    "кладовщик",
    "ричтрак",
    "штабелер",
    "штабл",
    "водитель",
    "разнорабоч",
    "оператор",
    "хаусмен",
    "боц",
    "слесарь",
    "сварщик",
    "монтажник",
    "бетонщик",
)


def has_demand(v):
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return float(v) > 0
    s = str(v).strip().lower()
    if s in ("да", "пс"):
        return True
    # e.g. "1 (1)", "5 (5)"
    m = re.search(r"\d+", s)
    if m:
        try:
            return float(m.group()) > 0
        except ValueError:
            pass
    try:
        return float(s.replace(",", ".")) > 0
    except ValueError:
        return False


def norm(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def parse_food_times(food_raw, food_star):
    text = f"{food_raw} {food_star}".lower().replace("ё", "е")
    m = re.search(r"([123])\s*(раз|раза)", text)
    if m:
        return int(m.group(1))
    if "три раз" in text:
        return 3
    if "два раз" in text:
        return 2
    if "один раз" in text or "1 раз" in text:
        return 1
    return None


def parse_ot(ot_raw):
    s = norm(ot_raw)
    if not s:
        return None
    try:
        return int(float(str(s).replace(",", ".").split()[0]))
    except ValueError:
        m = re.search(r"\d{1,3}", s)
        return int(m.group()) if m else None


def has_travel_compensation(comp_raw):
    s = norm(comp_raw)
    if not s:
        return False
    low = s.lower().replace("ё", "е")
    keys = ("проезд", "билет", "такси", "дорог", "транспорт", "ж/д", "жд", "авиа", "компенс")
    return any(k in low for k in keys)


def build_chips(food, food_times, housing, sb, sp_raw, travel_comp):
    chips = []
    if food:
        if food_times in (2, 3):
            chips.append({"text": f"питание {food_times} раза", "ok": True})
        else:
            chips.append({"text": "питание", "ok": True})
    if housing:
        chips.append({"text": "проживание", "ok": True})
    if sb in ("есть", "да"):
        chips.append({"text": "СБ есть", "ok": False})
    elif sb == "нет":
        chips.append({"text": "без СБ", "ok": False})

    sp_s = norm(sp_raw)
    if sp_s:
        if sp_s.lower() == "да":
            chips.append({"text": "семейные комнаты: уточните наличие", "ok": True})
        else:
            try:
                if float(str(sp_s).replace(",", ".")) > 0:
                    chips.append({"text": "семейные комнаты: есть", "ok": True})
            except ValueError:
                if sp_s.lower() not in ("нет", "0", "-"):
                    chips.append({"text": "семейные комнаты: уточните наличие", "ok": True})

    if travel_comp:
        chips.append({"text": "компенсация проезда", "ok": True})
    return chips


def split_job_lines(dol_raw):
    """Split Должность into role lines."""
    s = norm(dol_raw)
    if not s:
        return []
    if "\n" in s:
        parts = [p.strip() for p in s.splitlines() if p.strip()]
    else:
        # numbered inline: (1) a / (2) b
        if re.search(r"\(\d+\)", s) and "/" in s:
            parts = [p.strip() for p in re.split(r"\s*/\s*", s) if p.strip()]
        else:
            parts = [p.strip() for p in re.split(r"\s*/\s*", s) if p.strip()] if "/" in s else [s]
    return parts


def parse_job_line(line, line_idx):
    """Parse one должность line -> dict or None.

    Demand digit is required for multi-role expansion.
    Examples:
      (1) высотный штаблер (ричтрак) 1м
      комплектовщик 5ж
      горничная 5
      официант          -> count=None
    """
    raw = re.sub(r"\s+", " ", line).strip()
    if not raw:
        return None

    ordinal = None
    m_ord = re.match(r"^\((\d+)\)\s*", raw)
    if m_ord:
        ordinal = int(m_ord.group(1))
        raw = raw[m_ord.end() :].strip()

    count = None
    gender = None

    # trailing "5ж", "1м", "5 ж", "3М"
    m_tail = re.search(r"(?:^|\s)(\d+)\s*([мжМЖ])\s*$", raw)
    if m_tail:
        count = int(m_tail.group(1))
        gender = m_tail.group(2).lower()
        raw = raw[: m_tail.start()].strip()
    else:
        # trailing digit only: "горничная 5", "повар ... 3"
        m_num = re.search(r"(?:^|\s)(\d+)\s*$", raw)
        if m_num:
            count = int(m_num.group(1))
            raw = raw[: m_num.start()].strip()
        else:
            # gender-only suffix without count: "... ж" / "... м" — not open for search
            m_g = re.search(r"(?:^|\s)([мжМЖ])\s*$", raw)
            if m_g and len(raw) > 2:
                gender = m_g.group(1).lower()
                raw = raw[: m_g.start()].strip()

    name = raw.strip(" ·,-/")
    name = re.sub(r"\s+", " ", name).strip()
    if not name:
        return None

    if gender is None:
        low = name.lower().replace("ё", "е")
        if any(h in low for h in FEM_HINTS):
            gender = "ж"
        elif any(h in low for h in MASC_HINTS):
            gender = "м"

    return {
        "ordinal": ordinal if ordinal is not None else line_idx + 1,
        "line_idx": line_idx,
        "name": name,
        "name_key": name.lower().replace("ё", "е"),
        "count": count,
        "gender": gender,
    }


def parse_rates(rate_raw):
    """Map ordinal -> rate string; also list in order for positional match."""
    by_ord = {}
    ordered = []
    if rate_raw is None:
        return by_ord, ordered
    if isinstance(rate_raw, (int, float)):
        s = str(int(rate_raw)) if float(rate_raw) == int(rate_raw) else str(rate_raw)
        by_ord[1] = s
        ordered.append(s)
        return by_ord, ordered

    text = norm(rate_raw)
    if not text:
        return by_ord, ordered

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) == 1 and "/" in lines[0] and re.search(r"\d{3,5}", lines[0]):
        # rare: 3800/4500
        chunks = [c.strip() for c in lines[0].split("/") if c.strip()]
        if len(chunks) > 1:
            lines = chunks

    for i, ln in enumerate(lines, start=1):
        m = re.match(r"^(\d+)\s*[.)\-]?\s*(.+)$", ln)
        if m and re.search(r"\d{3,5}", m.group(2)):
            ord_n = int(m.group(1))
            rest = m.group(2).strip()
            num = re.search(r"\d{3,5}", rest)
            val = num.group(0) if num else rest.split()[0]
            by_ord[ord_n] = val
            ordered.append(val)
        else:
            num = re.search(r"\d{3,5}", ln)
            val = num.group(0) if num else ln[:20]
            by_ord[i] = val
            ordered.append(val)
    return by_ord, ordered


def pick_duty(duties_raw, job):
    """Try to take the duty block matching role ordinal / number."""
    text = norm(duties_raw)
    if not text:
        return ""
    lines = text.splitlines()
    n = job["ordinal"]
    # patterns: "1 ", "1)", "1.", "(1)"
    pat = re.compile(rf"^\s*\(?{n}\)?[.)\s:\-]")
    chunks = []
    capturing = False
    for ln in lines:
        if pat.match(ln):
            capturing = True
            chunks = [ln]
            continue
        if capturing:
            if re.match(r"^\s*\(?\d+\)?[.)\s:\-]", ln):
                break
            chunks.append(ln)
    if chunks:
        return "\n".join(chunks).strip()
    return text


def gender_flags(job, m_demand, zh_demand, sp_yes, sp_demand):
    """Card-level gender for search filters."""
    g = job.get("gender")
    if g == "м":
        return True, False
    if g == "ж":
        return False, True
    # unknown gender: inherit row demand / couples
    gm = bool(m_demand or sp_yes or sp_demand)
    gf = bool(zh_demand or sp_yes or sp_demand)
    if not gm and not gf:
        gm = gf = True
    return gm, gf


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["ПОТРЕБНОСТЬ1"]
    headers = [c.value for c in ws[1]]
    while headers and headers[-1] is None:
        headers.pop()

    zh_key = next((h for h in headers if h and str(h).strip() == "Ж"), None)

    tip_map = {
        "пищ": "пищ",
        "не пищ": "не пищ",
        "непищ": "не пищ",
        "склад": "склад",
        "отель": "отель",
        "стройка": "стройка",
        "тепл": "тепл",
        "теплицы": "тепл",
        "тт": "ТТ",
        "ферма": "ферма",
        "фер": "ферма",
    }
    tip_label = {
        "пищ": "пищевое",
        "не пищ": "не пищевое",
        "склад": "склад",
        "отель": "отель",
        "стройка": "стройка",
        "тепл": "теплицы",
        "ТТ": "торговая точка",
        "ферма": "ферма",
    }

    vacancies = []
    split_rows = 0
    legacy_rows = 0
    geo_cache = load_cache()
    geo_ok = 0
    geo_miss = 0
    use_network = "--geocode" in sys.argv

    for r in range(2, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        m = row.get("М")
        zh = row.get(zh_key) if zh_key else None
        sp = row.get("СП")
        if not (has_demand(m) or has_demand(zh) or has_demand(sp)):
            continue
        vid = row.get("ID")
        if vid is None:
            continue

        tip_raw = re.sub(r"\s+", " ", norm(row.get("Тип")).lower())
        tip_id = tip_map.get(tip_raw, tip_raw)
        sb = norm(row.get("СБ")).lower()
        med = norm(row.get("Мед книжка"))
        food_raw = norm(row.get("Питание"))
        food_star = norm(row.get("Питание*"))
        housing = norm(row.get("Проживание"))
        dol = norm(row.get("Должность"))
        for_max = norm(row.get("для МАКС"))
        comp = norm(row.get("Компенсации"))
        food_times = parse_food_times(food_raw, food_star)
        ot_num = parse_ot(row.get("ОТ"))
        travel_comp = has_travel_compensation(comp)
        duties_raw = norm(row.get("Обязанности"))

        try:
            id_num = int(float(vid)) if isinstance(vid, (int, float)) else str(vid).strip()
        except (TypeError, ValueError):
            id_num = str(vid).strip()

        job_lines = split_job_lines(dol)
        parsed = []
        for i, ln in enumerate(job_lines):
            p = parse_job_line(ln, i)
            if p:
                parsed.append(p)

        with_count = [p for p in parsed if p["count"] is not None and p["count"] > 0]
        # Expansion rule: if at least one role has a digit — only those roles become cards.
        # Legacy: no digits yet → one combined card so the object stays searchable.
        if with_count:
            roles = with_count
            expand = True
            split_rows += 1
        else:
            roles = [
                {
                    "ordinal": 1,
                    "line_idx": 0,
                    "name": " / ".join(p["name"] for p in parsed) if parsed else dol,
                    "name_key": " / ".join(p["name_key"] for p in parsed) if parsed else dol.lower(),
                    "count": None,
                    "gender": None,
                    "legacy": True,
                    "parts": parsed,
                }
            ]
            expand = False
            legacy_rows += 1

        rates_by_ord, rates_ordered = parse_rates(row.get("Ставка макс"))

        max_place = ""
        if for_max:
            lines = [ln.strip() for ln in for_max.splitlines() if ln.strip()]
            pin = next((ln for ln in lines if "📍" in ln), None)
            if not pin:
                pin = next(
                    (
                        ln
                        for ln in lines
                        if not ln.startswith("http")
                        and not ln.startswith("✅")
                        and not ln.startswith("🔥")
                    ),
                    "",
                )
            max_place = pin.replace("📍", "").strip()
            max_place = re.sub(r"https?://\S+", "", max_place).strip()
            max_place = re.sub(r"\s+", " ", max_place).strip(" ·,-")

        try:
            age_from = int(float(row.get("от") or 18))
        except (TypeError, ValueError):
            age_from = 18
        try:
            age_to = int(float(row.get("до") or 70))
        except (TypeError, ValueError):
            age_to = 70

        citizen_raw = norm(row.get("гр-во"))
        citizens = [c.strip() for c in re.split(r"[\n,;/]+", citizen_raw) if c.strip()]

        reg = norm(row.get("Регион")).replace(" область", "").replace("Область", "").strip()
        if reg in ("СПб", "СПБ"):
            reg = "Санкт-Петербург"

        photo = norm(row.get("фото"))
        if photo and not photo.startswith("http"):
            photo = ""

        food = food_raw.lower() in ("да", "есть", "1", "yes") or food_raw.lower().startswith("да")
        if food_times:
            food = True

        graph = norm(row.get("График вахты")).replace("\n", ", ")
        chips = build_chips(food, food_times, bool(housing), sb, sp, travel_comp)

        place_bits = [reg]
        if tip_label.get(tip_id):
            place_bits.append(tip_label[tip_id])
        if ot_num:
            place_bits.append(f"вахта от {ot_num}")

        sp_yes = str(sp).strip().lower() == "да" if sp is not None else False
        med_l = med.lower()
        m_demand = has_demand(m)
        zh_demand = has_demand(zh)
        sp_demand = has_demand(sp)

        for role in roles:
            if expand:
                job_title = role["name"]
                job_keys = [role["name_key"]]
                # rate: by ordinal first, else by original line index
                rate_s = rates_by_ord.get(role["ordinal"])
                if not rate_s and role["line_idx"] < len(rates_ordered):
                    rate_s = rates_ordered[role["line_idx"]]
                if not rate_s and rates_ordered:
                    rate_s = rates_ordered[0]
                rate_s = rate_s or ""
                duty = pick_duty(duties_raw, role)
                card_id = f"{id_num}-{role['ordinal']}"
                gm, gf = gender_flags(role, m_demand, zh_demand, sp_yes, sp_demand)
                demand_note = f"{role['count']} чел."
                if role.get("gender") == "м":
                    demand_note = f"{role['count']}М"
                elif role.get("gender") == "ж":
                    demand_note = f"{role['count']}Ж"
            else:
                job_title = role["name"]
                parts = role.get("parts") or []
                job_keys = [p["name_key"] for p in parts] if parts else [job_title.lower()]
                if rates_ordered:
                    rate_s = rates_ordered[0]
                else:
                    rate_s = ""
                duty = duties_raw
                card_id = str(id_num)
                gm = bool(m_demand or sp_yes or sp_demand)
                gf = bool(zh_demand or sp_yes or sp_demand)
                demand_note = " · ".join(
                    [
                        x
                        for x in [
                            f"М: {norm(m)}" if m_demand else "",
                            f"Ж: {norm(zh)}" if zh_demand else "",
                            f"СП: {norm(sp)}" if (sp_demand or sp_yes) else "",
                        ]
                        if x
                    ]
                )

            if job_title and max_place:
                place_part = max_place[0].lower() + max_place[1:] if len(max_place) > 1 else max_place.lower()
                title = f"{job_title} на {place_part}"
            elif max_place:
                title = max_place
            else:
                title = job_title
            title = re.sub(r"\s+", " ", title).strip(" ·,-")
            title = f"{title} (ID: {id_num})" if title else f"(ID: {id_num})"

            copy_parts = [f"ЦВЗ | {title}"]
            if expand:
                copy_parts.append(f"Потребность: {demand_note}")
            if reg:
                copy_parts.append(f"Регион: {reg}")
            if rate_s:
                copy_parts.append(f"Ставка: до {rate_s} руб/смена")
            if graph:
                copy_parts.append(f"График: {graph}")
            if food_raw:
                copy_parts.append(f"Питание: {food_raw}")
            if housing:
                copy_parts.append(f"Проживание: {housing.splitlines()[0]}")

            details = {
                "citizens": ", ".join(citizens),
                "age": f"{age_from}–{age_to}",
                "demand": demand_note,
                "sb": norm(row.get("СБ")),
                "sb_extra": norm(row.get("СБ*")),
                "med": med,
                "region": reg,
                "place": max_place,
                "type": tip_label.get(tip_id, tip_id),
                "ot": ot_num,
                "schedule": norm(row.get("График вахты")),
                "food": food_raw,
                "food_extra": food_star,
                "housing": housing,
                "settle": norm(row.get("Заселение")),
                "delivery": norm(row.get("Доставка")),
                "contract": norm(row.get("Оформление")),
                "clothes": norm(row.get("Спец одежда")),
                "compensation": comp,
                "jobs": job_title,
                "duties": duty if expand else duties_raw,
                "rate_extra": norm(row.get("Ставка*")),
                "extra": norm(row.get("доп инфа")),
                "neighbors": norm(row.get("Соседние регионы*")),
            }

            # Map pin: city/settlement only (never street / housing address)
            geo = geocode(max_place, reg, cache=geo_cache, use_network=use_network)
            lat = lng = None
            geo_label = ""
            if geo:
                lat, lng, geo_label = geo
                geo_ok += 1
            else:
                geo_miss += 1

            vacancies.append(
                {
                    "id": card_id,
                    "object_id": id_num,
                    "role_ordinal": role.get("ordinal"),
                    "title": title,
                    "place": " · ".join([b for b in place_bits if b]),
                    "pay": rate_s,
                    "photo": photo,
                    "chips": chips,
                    "duty": (duty or duties_raw)[:280],
                    "copy": "\n".join(copy_parts),
                    "region": reg,
                    "type": tip_id,
                    "food": food,
                    "food_times": food_times,
                    "housing": bool(housing),
                    "sb": sb,
                    "no_sb": sb == "нет",
                    "med": med,
                    "no_med": (not med) or ("не нуж" in med_l) or med_l == "нет",
                    "age_from": age_from,
                    "age_to": age_to,
                    "citizens": citizens,
                    "jobs": job_keys,
                    "gender_m": gm,
                    "gender_f": gf,
                    "ot": ot_num,
                    "short_shift": ot_num in (15, 21),
                    "travel_comp": travel_comp,
                    "sp": norm(sp),
                    "lat": lat,
                    "lng": lng,
                    "geo_label": geo_label,
                    "details": details,
                }
            )

    save_cache(geo_cache)
    payload = {"updated": "local-excel", "total": len(vacancies), "items": vacancies}
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"exported {len(vacancies)} cards -> {OUT}")
    print(f"rows split by role: {split_rows}; legacy combined: {legacy_rows}")
    print(f"geo pins: {geo_ok}; missing: {geo_miss}")
    print("short_shift", sum(1 for v in vacancies if v["short_shift"]))
    print("travel_comp", sum(1 for v in vacancies if v["travel_comp"]))


if __name__ == "__main__":
    main()
