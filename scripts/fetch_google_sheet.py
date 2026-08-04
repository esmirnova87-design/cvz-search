# -*- coding: utf-8 -*-
"""Download the ЦВЗ Google Sheet as xlsx for export_vacancies.py.

Uses a service account JSON (preferred, private sheet):
  secrets/google-service-account.json
  or GOOGLE_SERVICE_ACCOUNT_FILE / GOOGLE_APPLICATION_CREDENTIALS

Share the spreadsheet with client_email as Viewer.
Enable Google Sheets API in the GCP project.

Fallback: public link export if the sheet is shared «anyone with the link».
"""
from __future__ import annotations

import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "sheets-config.json"
OUT_XLSX = ROOT / "data" / "google-potrebnost.xlsx"
DEFAULT_SA = ROOT / "secrets" / "google-service-account.json"


def load_config():
    cfg = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    sid = cfg.get("spreadsheetId") or cfg.get("spreadsheet_id")
    if not sid:
        raise SystemExit("sheets-config.json: missing spreadsheetId")
    return cfg


def resolve_sa_path() -> Path | None:
    env = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE") or os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if env and Path(env).exists():
        return Path(env)
    if DEFAULT_SA.exists():
        return DEFAULT_SA
    return None


def download_public_xlsx(spreadsheet_id: str, dest: Path) -> None:
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "CVZ-sheet-sync/1.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        data = resp.read()
        ctype = (resp.headers.get("Content-Type") or "").lower()
    if len(data) < 1000 or "html" in ctype:
        raise RuntimeError(
            "Google вернул не Excel. Таблица закрыта для анонимного доступа."
        )
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)


def download_via_sheets_api(spreadsheet_id: str, dest: Path, sa_path: Path, preferred_sheet: str) -> None:
    """Read via Sheets API (no Drive API needed) and save as xlsx."""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        import openpyxl
    except ImportError as e:
        raise SystemExit(
            "Нужны пакеты: pip install openpyxl google-auth google-api-python-client\n" + str(e)
        ) from e

    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = service_account.Credentials.from_service_account_file(str(sa_path), scopes=scopes)
    service = build("sheets", "v4", credentials=creds, cache_discovery=False)

    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheets = meta.get("sheets", [])
    titles = [s["properties"]["title"] for s in sheets]
    if not titles:
        raise RuntimeError("В таблице нет листов")

    wb = openpyxl.Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    # Export all sheets so formulas/refs are less surprising; export uses ПОТРЕБНОСТЬ1
    for title in titles:
        result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=f"'{title}'", majorDimension="ROWS")
            .execute()
        )
        values = result.get("values", [])
        ws = wb.create_sheet(title=title[:31])
        for r_i, row in enumerate(values, start=1):
            for c_i, val in enumerate(row, start=1):
                ws.cell(r_i, c_i).value = val

    if preferred_sheet not in titles:
        print(f"warning: sheet '{preferred_sheet}' not in {titles}")

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    print(f"sheets via API: {len(titles)} tabs -> {dest}")


def fetch(dest: Path | None = None) -> Path:
    cfg = load_config()
    sid = cfg["spreadsheetId"]
    preferred = cfg.get("sheetName") or "ПОТРЕБНОСТЬ1"
    dest = dest or OUT_XLSX
    sa = resolve_sa_path()
    errors = []

    if sa:
        try:
            print(f"fetch via Sheets API + service account: {sa.name}")
            download_via_sheets_api(sid, dest, sa, preferred)
            print(f"saved {dest} ({dest.stat().st_size} bytes)")
            return dest
        except Exception as e:
            errors.append(f"sheets api: {e}")
            print(f"sheets api failed: {e}")

    try:
        print("fetch via public link export…")
        download_public_xlsx(sid, dest)
        print(f"saved {dest} ({dest.stat().st_size} bytes)")
        return dest
    except urllib.error.HTTPError as e:
        errors.append(f"public export HTTP {e.code}")
    except Exception as e:
        errors.append(f"public export: {e}")

    sa_email = ""
    if sa:
        try:
            sa_email = json.loads(sa.read_text(encoding="utf-8")).get("client_email", "")
        except Exception:
            pass

    msg = (
        "Не удалось скачать Google-таблицу.\n"
        + "\n".join(f"- {x}" for x in errors)
        + "\n\nПроверьте:\n"
        "1) В Google Cloud включён API: Google Sheets API\n"
        f"   https://console.developers.google.com/apis/library/sheets.googleapis.com\n"
        "2) В таблице выдан доступ Читатель на сервисный email:\n"
        f"   {sa_email or '(client_email из JSON)'}\n"
    )
    raise SystemExit(msg)


if __name__ == "__main__":
    fetch()
