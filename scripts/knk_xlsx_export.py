# -*- coding: utf-8 -*-
"""Export КНК xlsx (лист «Таблица») → paste text / apply plan helpers."""
from __future__ import annotations

import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

ROOT = Path(__file__).resolve().parents[1]


def parse_need_cell(val, vacancy: str = "", default_gender: str = "m") -> dict:
    """Return m, zh, has flags. 'по согласованию' → empty.
    default_gender: when value is bare number (row5→m, row6→zh in КНК Table).
    """
    m = zh = 0
    has_m = has_zh = False
    if val is None:
        return {"m": "", "zh": "", "has": False}
    if isinstance(val, (int, float)):
        n = int(val)
        if n <= 0:
            return {"m": "", "zh": "", "has": False}
        if default_gender == "zh":
            return {"m": "", "zh": str(n), "has": True}
        vac = (vacancy or "").lower()
        if re.search(r"упаковщиц|уборщиц|фасовщ|маркировщ", vac) and not re.search(
            r"грузчик|разнораб|комплект", vac
        ):
            return {"m": "", "zh": str(n), "has": True}
        return {"m": str(n), "zh": "", "has": True}

    text = str(val).strip()
    if not text or re.search(r"согласован", text, re.I):
        return {"m": "", "zh": "", "has": False}

    for n, g in re.findall(r"(\d+)\s*(муж\w*|жен\w*|м\b|ж\b)", text, flags=re.I):
        n = int(n)
        if n <= 0:
            continue
        if re.match(r"жен|ж\b", g, re.I):
            zh += n
            has_zh = True
        else:
            m += n
            has_m = True

    if not has_m and not has_zh:
        mm = re.match(r"^(\d+)\b", text.replace(" ", ""))
        if mm:
            n = int(mm.group(1))
            if n > 0:
                return parse_need_cell(n, vacancy, default_gender)

    return {
        "m": str(m) if has_m else "",
        "zh": str(zh) if has_zh else "",
        "has": has_m or has_zh,
    }


def rows_from_grid(grid: list[list]) -> list[dict]:
    """Parse Таблица-like grid: row0 project, row1 cities, row3 vacancy, row4-5 need."""
    if len(grid) < 6:
        return []
    # pad rows
    width = max(len(r) for r in grid[:6])
    def cell(r, c):
        row = grid[r] if r < len(grid) else []
        return row[c] if c < len(row) else None

    out = []
    for c in range(1, width):  # col0 = labels
        city = str(cell(1, c) or "").strip()
        if not city:
            continue
        vac = str(cell(3, c) or "").strip()
        a = parse_need_cell(cell(4, c), vac, default_gender="m")
        b = parse_need_cell(cell(5, c), vac, default_gender="zh")
        m = (int(a["m"] or 0) if a["m"] else 0) + (int(b["m"] or 0) if b["m"] else 0)
        zh = (int(a["zh"] or 0) if a["zh"] else 0) + (int(b["zh"] or 0) if b["zh"] else 0)
        has_m = bool(a["m"] or b["m"])
        has_zh = bool(a["zh"] or b["zh"])
        has = has_m or has_zh
        out.append(
            {
                "col": c + 1,  # 1-based like openpyxl for СПб col 4/5
                "city": city,
                "vacancy": vac.replace("\n", " ").strip(),
                "m": str(m) if has_m else "",
                "zh": str(zh) if has_zh else "",
                "sp": "да" if has_m and has_zh else "",
                "has": has,
            }
        )
    return out


def load_table_rows(xlsx: Path):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Таблица"]
    grid = []
    for r in range(1, 7):
        grid.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    return rows_from_grid(grid)


KNK_SHEET_ID = "18dIE1dyMFR41FCmgjdfMgUNe0aaIkebDtf9Cf1R8UOc"


def load_table_rows_from_google(spreadsheet_id: str = KNK_SHEET_ID) -> list[dict]:
    """Read лист «Таблица» via service account (no download needed)."""
    import json
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    sa = ROOT / "secrets" / "google-service-account.json"
    creds = service_account.Credentials.from_service_account_file(
        str(sa), scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
    )
    svc = build("sheets", "v4", credentials=creds, cache_discovery=False)
    data = (
        svc.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range="'Таблица'!A1:AC6")
        .execute()
        .get("values", [])
    )
    # normalize to 6 rows
    while len(data) < 6:
        data.append([])
    return rows_from_grid(data)


def to_paste_text(rows) -> str:
    lines = ["Город\tВакансия\tПотребность"]
    for r in rows:
        if r["has"]:
            need = []
            if r["m"]:
                need.append(f"{r['m']}М")
            if r["zh"]:
                need.append(f"{r['zh']}Ж")
            need_s = "/".join(need)
        else:
            need_s = "0"
        lines.append(
            "\t".join(
                [
                    r["city"].replace("\t", " "),
                    r["vacancy"].replace("\t", " ")[:80],
                    need_s,
                ]
            )
        )
    return "\n".join(lines)


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    use_google = "--google" in sys.argv or not args
    if use_google and (not args or args[0] in ("google", "--google")):
        print("source: Google Sheet", KNK_SHEET_ID)
        rows = load_table_rows_from_google()
    else:
        xlsx = Path(args[0]) if args else Path(r"c:\Users\user\Downloads\Заявка КНК.xlsx")
        print("source: file", xlsx)
        rows = load_table_rows(xlsx)
    active = [r for r in rows if r["has"]]
    print(f"cols={len(rows)} active={len(active)}")
    for r in active:
        print(f"  {r['m'] or '0'}М/{r['zh'] or '0'}Ж | {r['city'][:40]} | {r['vacancy'][:40]}")
    paste = to_paste_text(rows)
    out = ROOT / "scripts" / "_knk_paste.tsv"
    out.write_text(paste, encoding="utf-8")
    print("paste ->", out)


if __name__ == "__main__":
    main()
