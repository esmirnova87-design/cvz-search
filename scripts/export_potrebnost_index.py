# -*- coding: utf-8 -*-
"""Export compact index of ПОТРЕБНОСТЬ1 for admin need-update matching."""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "google-potrebnost.xlsx"
OUT = ROOT / "potrebnost-index.json"


def norm_num(v) -> str:
    if v is None or v == "":
        return ""
    try:
        f = float(str(v).replace(",", "."))
        if f == int(f):
            return str(int(f))
        return str(v).strip()
    except ValueError:
        return str(v).strip()


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing {XLSX}; run fetch_google_sheet.py first")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["ПОТРЕБНОСТЬ1"]
    headers = [
        (str(c.value).strip() if c.value is not None else "")
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    idx = {h: i for i, h in enumerate(headers) if h}

    def cell(row, name):
        i = idx.get(name)
        return None if i is None else row[i]

    rows = []
    for r_i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cust = cell(r, "Заказчик")
        oid = cell(r, "ID")
        if not cust or oid is None or oid == "":
            continue

        def s(v):
            return "" if v is None else str(v).strip().replace("\r", "")

        sp_raw = cell(r, "СП")
        sp = norm_num(sp_raw)
        if not sp and sp_raw is not None and str(sp_raw).strip():
            sp = str(sp_raw).strip().lower()

        rows.append(
            {
                "id": norm_num(oid) or s(oid),
                "customer": s(cust),
                "object": s(cell(r, "Объект")),
                "m": norm_num(cell(r, "М")),
                "zh": norm_num(cell(r, "Ж")),
                "sp": sp,
                "job": s(cell(r, "Должность")),
                "rate": s(cell(r, "Ставка макс")),
                "max": s(cell(r, "МАКС")),
                "sheetRow": r_i,
            }
        )

    OUT.write_text(
        json.dumps({"updated": "from-xlsx", "rows": rows}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"wrote {OUT} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
