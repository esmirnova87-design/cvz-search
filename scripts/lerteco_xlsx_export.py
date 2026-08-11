# -*- coding: utf-8 -*-
"""Export Lerteco xlsx (лист 2026 заявка, Вахта=да) → text for site parser / apply plan."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

ROOT = Path(__file__).resolve().parents[1]


def parse_need_cell(need, requirements: str = "", job: str = "") -> dict:
    text = str(need if need is not None else "").strip()
    req = str(requirements or "")
    job_l = str(job or "").lower()

    m = zh = 0
    has_m = has_zh = False

    # «2 мужчин», «40 мужчин» — «0 мужчин/женщин» не считаем потребностью
    for n, g in re.findall(r"(\d+)\s*(мужчин\w*|муж\b|жен\w*|девуш\w*)", text, flags=re.I):
        n = int(n)
        if n <= 0:
            continue
        if re.match(r"жен|девуш", g, re.I):
            zh += n
            has_zh = True
        else:
            m += n
            has_m = True

    if not has_m and not has_zh:
        num = None
        if isinstance(need, (int, float)) and float(need) > 0:
            num = int(need)
        else:
            mm = re.match(r"^(\d+(?:[.,]\d+)?)\b", text.replace(" ", ""))
            if mm:
                try:
                    num = int(float(mm.group(1).replace(",", ".")))
                except ValueError:
                    num = None
        if num and num > 0:
            # пол из требований / должности
            if re.search(r"женск|женщин|девуш", req + " " + job_l, re.I) and not re.search(
                r"мужск|мужчин", req, re.I
            ):
                zh, has_zh = num, True
            elif re.search(r"мужск|мужчин|муж\b", req, re.I) or re.search(
                r"грузчик|водитель|оператор|сборщик|комплектов|слесарь|свар|маляр|вап|вэш",
                job_l,
                re.I,
            ):
                m, has_m = num, True
            else:
                # по умолчанию М (у Lerteco чаще мужские роли)
                m, has_m = num, True

    # стикеровщицы / упаковщицы в должности без числа пола
    if not has_m and not has_zh and re.search(r"стикеровщиц|упаковщиц|уборщиц", job_l):
        pass  # need was 0

    sp = ""
    if has_m and has_zh:
        sp = "да"
    return {
        "m": str(m) if has_m else "",
        "zh": str(zh) if has_zh else "",
        "sp": sp,
        "has": has_m or has_zh,
        "raw_need": text[:80],
    }


def load_vahta_rows_from_values(rows_2d: list[list], start_row: int = 2) -> list[dict]:
    """Parse rows from Sheets API / openpyxl grid (0-based list of lists, col A = index 0)."""
    out = []
    for i, row in enumerate(rows_2d):
        def cell(idx):
            return row[idx] if idx < len(row) else None

        vahta = str(cell(4) or "").strip().lower()  # E
        if vahta != "да":
            continue
        obj = str(cell(0) or "").strip()
        job = str(cell(1) or "").strip()
        need = cell(2)
        req = str(cell(10) or "")  # K
        addr = str(cell(7) or "").strip()  # H
        if not obj:
            continue
        parsed = parse_need_cell(need, req, job)
        out.append(
            {
                "excelRow": start_row + i,
                "object": obj,
                "job": job,
                "addr": addr,
                "need": need,
                **parsed,
            }
        )
    return out


def load_vahta_rows(xlsx: Path):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["2026 заявка"]
    grid = []
    for r in range(2, ws.max_row + 1):
        grid.append([ws.cell(r, c).value for c in range(1, 12)])
    return load_vahta_rows_from_values(grid, start_row=2)


LERTECO_SHEET_ID = "1NAgklIGANNrMNTeUkz8iRsNij9Cu91SVc5bXgEuwboc"


def load_vahta_rows_from_google(spreadsheet_id: str = LERTECO_SHEET_ID) -> list[dict]:
    """Read лист «2026 заявка» via service account (Вахта=да)."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    sa = ROOT / "secrets" / "google-service-account.json"
    creds = service_account.Credentials.from_service_account_file(
        str(sa), scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
    )
    svc = build("sheets", "v4", credentials=creds, cache_discovery=False)
    data = (
        svc.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range="'2026 заявка'!A2:K")
        .execute()
        .get("values", [])
    )
    return load_vahta_rows_from_values(data, start_row=2)


def to_paste_text(rows) -> str:
    """TSV for site parseLerteco."""
    lines = ["Название объекта\tДолжность\tПотребность\tВахта\tАдрес объекта\tТребования"]
    for x in rows:
        # only rows with need>0 for paste of "active"; zeros still useful for clear logic if full dump
        need_out = x["raw_need"] or (x["m"] + "М" if x["m"] else "") or (x["zh"] + "Ж" if x["zh"] else "0")
        if x["m"] and not re.search(r"муж|М", need_out, re.I):
            need_out = f"{x['m']} мужчин"
        if x["zh"] and not re.search(r"жен|Ж", need_out, re.I):
            need_out = (need_out + " " if need_out and need_out != "0" else "") + f"{x['zh']} женщин"
        lines.append(
            "\t".join(
                [
                    x["object"].replace("\n", " ").replace("\t", " "),
                    x["job"].replace("\n", " ").replace("\t", " ")[:80],
                    need_out.replace("\n", " ").replace("\t", " "),
                    "да",
                    x["addr"].replace("\n", " ").replace("\t", " ")[:80],
                    "",
                ]
            )
        )
    return "\n".join(lines)


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    use_file = bool(args) and args[0].endswith((".xlsx", ".xlsm"))
    if use_file:
        print("source: file", args[0])
        rows = load_vahta_rows(Path(args[0]))
    else:
        print("source: Google Sheet", LERTECO_SHEET_ID)
        rows = load_vahta_rows_from_google()
    active = [r for r in rows if r["has"]]
    print(f"vahta=да: {len(rows)}, with need: {len(active)}")
    for r in active:
        print(
            f"  {r['m'] or '0'}М/{r['zh'] or '0'}Ж | {r['object'][:50].replace(chr(10),' ')} | {r['job'][:40].replace(chr(10),' ')}"
        )

    paste = to_paste_text(rows)  # full vahta dump for clear logic
    out_txt = ROOT / "scripts" / "_lerteco_paste.tsv"
    out_txt.write_text(paste, encoding="utf-8")
    print("paste ->", out_txt)


if __name__ == "__main__":
    main()
